from bs4 import BeautifulSoup  # импортируем библиотеку BeautifulSoup, библиотека для синтаксического разбора файлов HTML/XML, которая может 
# преобразовать даже неправильную разметку в дерево синтаксического разбора.
import requests  # импортируем библиотеку requests, позволяет очень легко отправлять HTTP/1.1 запросы. 
from openpyxl import load_workbook


def parse1():
    fn = 'лаба1.xlsx'
    wb = load_workbook(fn) # Функция load_workbook () принимает имя файла в качестве аргумента и возвращает объект рабочей книги, который представляет файл.
    ws = wb['данные'] # Рабочий лист можно получить, используя его имя в качестве ключа экземпляра созданной книги Excel
    description1 = []
    description2 = []
    description3 = []
    count=2
    s='https://www.chitai-gorod.ru/catalog/collections/bestsell?page='#адрес страницы
    for i in range(1, count + 1):
        url1 = s + str(i)  # передаем необходимы URL адрес
        page1 = requests.get(url1)  # отправляем запрос методом Get на данный адрес и получаем ответ в переменную
        #Метод GET указывает на то, что происходит попытка извлечь данные из определенного ресурса. Для того, чтобы выполнить запрос GET, используется requests.get().
        print(page1.status_code)  # смотрим ответ, можно увидеть код состояния, который возвращается с сервера
        soup1 = BeautifulSoup(page1.text, "html.parser")  # передаем страницу в bs4, разбираем документ и передаем в конструктор BeautifulSoup 
        # Этот объект принимает в качестве аргументов документ из Requests (содержимое ответа сервера), а затем анализирует его
        #указав в кавычках как он нам поможет 'html.parcer'
        block1 = soup1.findAll('article', class_='product-card product-card product')  # находим контейнер с нужным классом


        for data in block1:  # проходим циклом по содержимому контейнера
            nazv = data.find(class_='product-title__head')
            author = data.find(class_='product-title__author')
            price = data.find(class_='product-price__value')
            if ((nazv and author and price) is not None):
                description1.append(nazv.text)
                description2.append(author.text)
                description3.append(price.text)


    print(description1, description2, description3)
    ech = 1
    for elem1, elem2, elem3 in zip(description1, description2, description3): # Функция zip() в Python создает итератор, который объединяет 
        #элементы из нескольких источников данных.
        cell = ws.cell(1, ech) # получить объект-ячейку - команда cell, вписываем в ячейку первой строки 1 столбца
        cell.value = elem1

        cell = ws.cell(2, ech)
        cell.value = elem2

        cell = ws.cell(3, ech)
        cell.value = elem3

        ech += 1

    wb.save(fn)#сохраняю изменения в фаиле
    wb.close()#закрываю фаил




